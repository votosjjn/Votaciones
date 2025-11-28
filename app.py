import os
import json
import datetime
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_from_directory
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "cambiame_ya_123")

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
os.makedirs(DATA_DIR, exist_ok=True)

# file paths
USERS_FILE = os.path.join(DATA_DIR, "usuarios.json")
CAND_PRIM = os.path.join(DATA_DIR, "candidatos_primaria.json")
CAND_SEC = os.path.join(DATA_DIR, "candidatos_secundaria.json")
VOTES_XLSX = os.path.join(DATA_DIR, "votos.xlsx")

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "jjnkvotos2025")

# -------- Helpers JSON --------
def load_json(path, default):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        return default
    with open(path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except Exception:
            return default

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# Ensure default JSON exists
load_json(USERS_FILE, {})
load_json(CAND_PRIM, [
    {"id": "p1", "nombre": "Miguel Amaya", "foto": "/static/img/candidato1.png"},
    {"id": "p2", "nombre": "Sofia Castañeda", "foto": "/static/img/candidato2.png"},
    {"id": "p3", "nombre": "Johan Ferney", "foto": "/static/img/candidato3.png"}
])
load_json(CAND_SEC, [
    {"id": "s1", "nombre": "Miguel Amaya", "foto": "/static/img/candidato1.png"},
    {"id": "s2", "nombre": "Sofia Castañeda", "foto": "/static/img/candidato2.png"},
    {"id": "s3", "nombre": "Johan Ferney", "foto": "/static/img/candidato3.png"}
])

# ---------- Excel init ----------
def ensure_excel():
    if not os.path.exists(VOTES_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "votos"
        ws.append(["timestamp", "documento", "nivel", "candidato_id", "candidato_nombre"])
        wb.save(VOTES_XLSX)

ensure_excel()

# ========== RUTAS ==========

@app.route("/")
def menu():
    return render_template("menu.html")

@app.route("/gracias")
def gracias():
    return render_template("gracias.html")

# DOCUMENTO LOGIN
@app.route("/documento/<nivel>", methods=["GET", "POST"])
def documento(nivel):
    nivel = nivel.lower()
    if nivel not in ("primaria", "secundaria"):
        return redirect(url_for("menu"))
    
    if request.method == "POST":
        doc = request.form.get("doc", "").strip()
        if not doc:
            return render_template("documento.html", nivel=nivel, error="Ingresa el número de documento")
        
        users = load_json(USERS_FILE, {})
        if doc not in users:
            return render_template("error.html", mensaje="Documento no registrado. Contacta al administrador.")
        if not users[doc].get("vino", False):
            return render_template("error.html", mensaje="No estás marcado como presente. No puedes votar.")

        users[doc]["presente"] = True
        save_json(USERS_FILE, users)
        return redirect(url_for("candidatos", nivel=nivel, doc=doc))
    
    return render_template("documento.html", nivel=nivel)

# MOSTRAR CANDIDATOS
@app.route("/candidatos/<nivel>")
def candidatos(nivel):
    nivel = nivel.lower()
    doc = request.args.get("doc", "").strip()
    if nivel not in ("primaria", "secundaria") or not doc:
        return redirect(url_for("menu"))
    
    users = load_json(USERS_FILE, {})
    if doc not in users:
        return redirect(url_for("documento", nivel=nivel))
    
    candidatos = load_json(CAND_PRIM if nivel == "primaria" else CAND_SEC, [])
    return render_template("candidatos.html", nivel=nivel, candidatos=candidatos, doc=doc)

# VOTAR (AJAX)
@app.route("/vote/<nivel>", methods=["POST"])
def vote(nivel):
    nivel = nivel.lower()
    data = request.get_json() or {}
    doc = str(data.get("doc", "")).strip()
    cid = data.get("candidato")

    if nivel not in ("primaria", "secundaria"):
        return jsonify({"ok": False, "msg": "Nivel inválido"}), 400
    if not doc or not cid:
        return jsonify({"ok": False, "msg": "Faltan datos"}), 400
    
    users = load_json(USERS_FILE, {})
    if doc not in users:
        return jsonify({"ok": False, "msg": "Documento no registrado"}), 403
    if not users[doc].get("vino", False) or not users[doc].get("presente", False):
        return jsonify({"ok": False, "msg": "No estás autorizado para votar"}), 403

    key = "voto_primaria" if nivel == "primaria" else "voto_secundaria"
    if users[doc].get(key):
        return jsonify({"ok": False, "msg": "Ya votaste"}), 403

    cand_list = load_json(CAND_PRIM if nivel == "primaria" else CAND_SEC, [])
    cand_name = None
    for c in cand_list:
        if c["id"] == cid:
            cand_name = c.get("nombre", "")
            break
    if cand_name is None:
        return jsonify({"ok": False, "msg": "Candidato inválido"}), 400

    users[doc][key] = cid
    save_json(USERS_FILE, users)

    ensure_excel()
    wb = load_workbook(VOTES_XLSX)
    ws = wb.active
    ts = datetime.datetime.utcnow().isoformat()
    ws.append([ts, doc, nivel, cid, cand_name])
    wb.save(VOTES_XLSX)

    return jsonify({"ok": True})

# ----------- RUTAS ADMIN -----------

def count_votes(nivel):
    """
    Lee votos desde votos.xlsx y devuelve:
      - lista de candidatos del nivel (dicts)
      - conteo {candidato_id: total}
    """
    ensure_excel()
    wb = load_workbook(VOTES_XLSX)
    ws = wb.active

    conteo = {}
    # cada fila: timestamp, documento, nivel, candidato_id, candidato_nombre
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        ts, doc, lvl, cid, cname = row
        if lvl == nivel:
            conteo[cid] = conteo.get(cid, 0) + 1

    cand_list = load_json(CAND_PRIM if nivel == "primaria" else CAND_SEC, [])
    return cand_list, conteo

@app.route("/admin", methods=["GET","POST"])
def admin_login():
    if request.method=="POST":
        pwd = request.form.get("pwd","")
        if pwd==ADMIN_PASSWORD:
            session["admin"]=True
            return redirect(url_for("admin_panel"))
        return render_template("admin_login.html", error="Contraseña incorrecta")
    return render_template("admin_login.html")

@app.route("/logout_admin")
def logout_admin():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))

@app.route("/admin/panel")
def admin_panel():

    if not session.get("admin"):
        return redirect(url_for("admin_login"))

    cand_p, c_p = count_votes("primaria")
    cand_s, c_s = count_votes("secundaria")
    users = load_json(USERS_FILE, {})

    for c in cand_p:
        c_p.setdefault(c["id"], 0)
    for c in cand_s:
        c_s.setdefault(c["id"], 0)

    total_prim = sum(c_p.values()) if c_p else 0
    total_sec = sum(c_s.values()) if c_s else 0

    porcentajes_prim = {}
    porcentajes_sec = {}
    for cid, cnt in c_p.items():
        porcentajes_prim[cid] = round(cnt / total_prim * 100, 1) if total_prim > 0 else 0
    for cid, cnt in c_s.items():
        porcentajes_sec[cid] = round(cnt / total_sec * 100, 1) if total_sec > 0 else 0

    # ------- COMBINADO AGRUPADO POR NOMBRE -------
    combinado = {}

    # Unir listas de candidatos
    todos = cand_p + cand_s

    for c in todos:
        nombre = c["nombre"]
        cid = c["id"]
        votos = c_p.get(cid, 0) + c_s.get(cid, 0)
        combinado[nombre] = votos

    # Ordenar por nombre (diccionario ordenado)
    combinado = dict(sorted(combinado.items(), key=lambda x: x[0].lower()))
    # ----------------------------------------------

    return render_template(
        "admin_panel.html",
        candidatos_prim=cand_p,
        conteo_prim=c_p,
        total_prim=total_prim,
        porcentajes_prim=porcentajes_prim,
        candidatos_sec=cand_s,
        conteo_sec=c_s,
        total_sec=total_sec,
        porcentajes_sec=porcentajes_sec,
        combinado=combinado,
        usuarios=users
    )



# --- cargar usuarios desde excel ---
UPLOAD_FOLDER = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
ALLOWED_EXT = {"xlsx"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXT

@app.route("/admin/cargar_usuarios", methods=["GET","POST"])
def cargar_usuarios():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    if request.method=="POST":
        file = request.files.get("archivo")
        if not file or file.filename=="":
            return render_template("cargar_usuarios.html", error="Selecciona un archivo .xlsx")
        if not allowed_file(file.filename):
            return render_template("cargar_usuarios.html", error="El archivo debe ser .xlsx")
        filename = secure_filename(file.filename)
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(path)

        wb = load_workbook(path)
        ws = wb.active
        usuarios = load_json(USERS_FILE, {})
        for row in ws.iter_rows(min_row=2, values_only=True):
            curso, documento, nombre, vino = row
            if not documento:
                continue
            documento = str(documento).strip()
            usuarios[documento] = {
                "curso": curso,
                "nombre": nombre,
                "vino": str(vino).lower() in ("si","sí","true","1"),
                "presente": False
            }
        save_json(USERS_FILE, usuarios)
        return render_template("cargar_usuarios.html", ok="Usuarios cargados correctamente.")
    return render_template("cargar_usuarios.html")

# --- marcar presente/ausente ---
@app.route("/admin/marcar_presente/<doc>", methods=["POST"])
def marcar_presente(doc):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    users = load_json(USERS_FILE, {})
    if doc in users:
        users[doc]["presente"]=True
        save_json(USERS_FILE, users)
    return redirect(url_for("admin_panel"))

@app.route("/admin/marcar_ausente/<doc>", methods=["POST"])
def marcar_ausente(doc):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    users = load_json(USERS_FILE, {})
    if doc in users:
        users[doc]["presente"]=False
        save_json(USERS_FILE, users)
    return redirect(url_for("admin_panel"))

# --- borrar votos ---
@app.route("/admin/borrar_votos/<nivel>", methods=["POST"])
def borrar_votos(nivel):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    nivel = nivel.lower()
    if nivel not in ("primaria","secundaria"):
        return redirect(url_for("admin_panel"))
    users = load_json(USERS_FILE,{})
    key = "voto_primaria" if nivel=="primaria" else "voto_secundaria"
    for u in users.values():
        if key in u:
            del u[key]
    save_json(USERS_FILE, users)
    wb = load_workbook(VOTES_XLSX)
    ws = wb.active
    # Safeguard: solo eliminar si hay filas
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(VOTES_XLSX)
    return redirect(url_for("admin_panel"))

# Static
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(os.path.join(BASE, "static"), filename)

@app.route("/borrar_votos_primaria", methods=["POST"])
def borrar_votos_primaria():
    global votos_primaria
    votos_primaria = {"p1": 0, "p2": 0, "p3": 0}
    return redirect(url_for("admin_panel"))


@app.route("/borrar_votos_secundaria", methods=["POST"])
def borrar_votos_secundaria():
    global votos_secundaria
    votos_secundaria = {"s1": 0, "s2": 0, "s3": 0}
    return redirect(url_for("admin_panel"))


@app.route("/error")
def error():
    mensaje = request.args.get("mensaje", "Ocurrió un error desconocido.")
    return render_template("error.html", mensaje=mensaje)


if __name__=="__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

